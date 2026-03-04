import json
import threading
import webbrowser
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook


HOST = "127.0.0.1"
PORT = 8000
BASE_DIR = Path(__file__).resolve().parent
CONTACTS_XLSX = BASE_DIR / "contacts.xlsx"

STORE_DATA = {
    "company": {
        "name": "Công ty TNHH Thương Mại và Dịch vụ TH-Ecohome",
        "short_name": "TH-Ecohome",
        "tagline": "Thiết kế, thi công nhà phố, nội thất và cải tạo công trình theo phong cách hiện đại",
        "address": "25/5 ấp Bến Lức 6, Xã Bến Lức, Tỉnh Tây Ninh",
        "phone": "0367954406",
        "zalo": "0367954406",
        "hours": "Tư vấn mỗi ngày: 08:00 - 20:00",
        "website_note": "Đơn vị chuyên thiết kế, thi công công trình dân dụng, nội thất và cải tạo không gian sống.",
    },
    "hero_slides": [
        {
            "eyebrow": "TH-Ecohome",
            "title": "Thiết kế và thi công nhà phố tối giản, trẻ trung, đúng công năng",
            "description": "Đồng hành từ ý tưởng, bố trí công năng, phối cảnh 3D đến thi công hoàn thiện cho công trình dân dụng.",
            "image": "https://images.unsplash.com/photo-1600585154526-990dced4db0d?auto=format&fit=crop&w=1400&q=80",
        },
        {
            "eyebrow": "Nội thất và hoàn thiện",
            "title": "Không gian đẹp, sạch, tối ưu chi phí cho nhà ở và căn hộ",
            "description": "Tập trung vào vật liệu bền, tông màu hiện đại và giải pháp nội thất phù hợp nhu cầu sinh hoạt thực tế.",
            "image": "https://images.unsplash.com/photo-1600210492493-0946911123ea?auto=format&fit=crop&w=1400&q=80",
        },
        {
            "eyebrow": "Cải tạo công trình",
            "title": "Làm mới nhà cũ, cải tạo mặt tiền, bếp, phòng tắm và không gian kinh doanh",
            "description": "Thi công gọn gàng, rõ ràng từng hạng mục, ưu tiên tiến độ và khả năng đưa vào sử dụng nhanh.",
            "image": "https://images.unsplash.com/photo-1505693416388-ac5ce068fe85?auto=format&fit=crop&w=1400&q=80",
        },
    ],
    "highlights": [
        {"title": "Thiết kế nhà phố", "text": "Giải pháp mặt bằng, mặt tiền và nội thất thống nhất."},
        {"title": "Thi công trọn gói", "text": "Theo sát tiến độ, quản lý hạng mục và chất lượng hoàn thiện."},
        {"title": "Nội thất hiện đại", "text": "Tối giản, trẻ trung, dễ sử dụng và dễ bảo trì."},
        {"title": "Cải tạo linh hoạt", "text": "Nâng cấp công năng cho nhà ở, căn hộ và công trình nhỏ."},
    ],
    "services": [
        {"name": "Thiết kế kiến trúc nhà phố", "summary": "Mặt bằng công năng, mặt tiền, phối cảnh và định hướng vật liệu đồng bộ.", "image": "https://images.unsplash.com/photo-1600607687939-ce8a6c25118c?auto=format&fit=crop&w=900&q=80"},
        {"name": "Thi công nhà dân dụng", "summary": "Thi công phần thô, hoàn thiện và phối hợp từng hạng mục theo kế hoạch.", "image": "https://images.unsplash.com/photo-1600585154340-be6161a56a0c?auto=format&fit=crop&w=900&q=80"},
        {"name": "Thiết kế và thi công nội thất", "summary": "Không gian phòng khách, bếp, phòng ngủ, phòng tắm theo phong cách hiện đại.", "image": "https://images.unsplash.com/photo-1600566753190-17f0baa2a6c3?auto=format&fit=crop&w=900&q=80"},
        {"name": "Cải tạo công trình", "summary": "Sửa chữa, nâng cấp, thay đổi công năng và làm mới không gian sử dụng.", "image": "https://images.unsplash.com/photo-1600607687644-c7171b42498f?auto=format&fit=crop&w=900&q=80"},
    ],
    "project_types": [
        {"name": "Nhà phố dân dụng", "style": "Mặt tiền gọn, sáng, nhấn mạnh tỷ lệ và ánh sáng", "image": "https://images.unsplash.com/photo-1600047509807-ba8f99d2cdde?auto=format&fit=crop&w=900&q=80"},
        {"name": "Nội thất căn hộ", "style": "Tối giản, trẻ trung, tối ưu diện tích", "image": "https://images.unsplash.com/photo-1494526585095-c41746248156?auto=format&fit=crop&w=900&q=80"},
        {"name": "Phòng bếp và phòng tắm", "style": "Ưu tiên công năng, vật liệu bền và dễ vệ sinh", "image": "https://images.unsplash.com/photo-1556911220-bff31c812dba?auto=format&fit=crop&w=900&q=80"},
        {"name": "Cải tạo mặt tiền", "style": "Làm mới hình ảnh công trình nhanh gọn, hiệu quả", "image": "https://images.unsplash.com/photo-1511818966892-d7d671e672a2?auto=format&fit=crop&w=900&q=80"},
    ],
    "projects": [
        {"name": "Nhà phố 3 tầng phong cách hiện đại", "tag": "Thiết kế + Thi công", "text": "Tối ưu thông gió, ánh sáng tự nhiên và khu vực sinh hoạt chung cho gia đình trẻ.", "image": "https://images.unsplash.com/photo-1605146769289-440113cc3d00?auto=format&fit=crop&w=900&q=80"},
        {"name": "Căn hộ 2 phòng ngủ tối giản", "tag": "Nội thất", "text": "Bố trí lại phòng khách - bếp liên thông, sử dụng tông gỗ sáng và bề mặt trung tính.", "image": "https://images.unsplash.com/photo-1505693416388-ac5ce068fe85?auto=format&fit=crop&w=900&q=80"},
        {"name": "Cải tạo bếp và phòng tắm", "tag": "Cải tạo", "text": "Nâng cấp vật liệu, đèn, thiết bị và giải pháp lưu trữ gọn gàng cho nhà đã ở lâu năm.", "image": "https://images.unsplash.com/photo-1484154218962-a197022b5858?auto=format&fit=crop&w=900&q=80"},
    ],
    "process": [
        {"step": "01", "title": "Tiếp nhận nhu cầu", "text": "Trao đổi mục tiêu, ngân sách, phong cách và hiện trạng công trình."},
        {"step": "02", "title": "Lên ý tưởng", "text": "Đề xuất hướng thiết kế, mặt bằng và bộ vật liệu phù hợp."},
        {"step": "03", "title": "Triển khai thi công", "text": "Chốt hạng mục, lập tiến độ và theo dõi từng giai đoạn thi công."},
        {"step": "04", "title": "Bàn giao - hỗ trợ", "text": "Nghiệm thu, bàn giao và tiếp tục đồng hành khi công trình đưa vào sử dụng."},
    ],
    "articles": [
        {"title": "3 lưu ý khi thiết kế nhà phố cho gia đình trẻ", "date": "03/03/2026", "excerpt": "Cân bằng giữa công năng, thông gió, ánh sáng và khả năng thay đổi theo nhu cầu sử dụng.", "image": "https://images.unsplash.com/photo-1511818966892-d7d671e672a2?auto=format&fit=crop&w=900&q=80"},
        {"title": "Cách cải tạo phòng bếp đẹp mà vẫn gọn sạch", "date": "28/02/2026", "excerpt": "Tập trung vào bố trí giao thông, bề mặt dễ lau chùi và hệ tủ lưu trữ hợp lý.", "image": "https://images.unsplash.com/photo-1556911220-bff31c812dba?auto=format&fit=crop&w=900&q=80"},
        {"title": "Phong cách nội thất tối giản phù hợp nhà ở hiện đại", "date": "23/02/2026", "excerpt": "Dùng màu sắc tiết chế, vật liệu bền và ánh sáng để tạo cảm giác rộng và thoáng.", "image": "https://images.unsplash.com/photo-1494526585095-c41746248156?auto=format&fit=crop&w=900&q=80"},
    ],
    "testimonials": [
        {"name": "Chị Nhi", "role": "Chủ nhà", "quote": "Bên TH-Ecohome làm việc rõ ràng, dễ trao đổi và phối hợp rất nhanh trong quá trình cải tạo nhà."},
        {"name": "Anh Thanh", "role": "Khách hàng nội thất", "quote": "Phong cách trẻ trung, hiện đại và giải pháp nội thất rất vừa tầm sinh hoạt của gia đình mình."},
        {"name": "Chị Trúc", "role": "Khách hàng nhà phố", "quote": "Từ khâu tư vấn đến thi công đều gọn gàng, dễ theo dõi và không bị rối thông tin."},
    ],
}


def ensure_contacts_workbook() -> None:
    if CONTACTS_XLSX.exists():
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Contacts"
    sheet.append(
        [
            "Submitted At",
            "Form Type",
            "Full Name",
            "Phone",
            "Message",
            "Language",
            "Source",
        ]
    )
    workbook.save(CONTACTS_XLSX)


def save_contact(payload: dict) -> None:
    ensure_contacts_workbook()
    workbook = load_workbook(CONTACTS_XLSX)
    sheet = workbook.active
    sheet.append(
        [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            payload.get("form_type", ""),
            payload.get("name", ""),
            payload.get("phone", ""),
            payload.get("message", ""),
            payload.get("language", ""),
            payload.get("source", ""),
        ]
    )
    workbook.save(CONTACTS_XLSX)

HTML_PAGE = """<!DOCTYPE html>
<html lang="vi">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>TH-Ecohome | Thiết kế, thi công, nội thất và cải tạo công trình</title>
  <meta name="description" content="Công ty TNHH Thương Mại và Dịch vụ TH-Ecohome chuyên thiết kế, thi công nhà phố dân dụng, nội thất và cải tạo công trình." />
  <link rel="preconnect" href="https://fonts.googleapis.com" />
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin />
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet" />
  <style>
    :root{--bg:#f4f5f7;--surface:#ffffff;--surface-soft:#f8fafc;--ink:#132033;--muted:#5f6b7a;--line:#d9dee7;--brand:#f28b20;--brand-dark:#cf6c08;--green:#17365d;--shadow:0 18px 38px rgba(19,32,51,.08);--container:min(1180px,calc(100vw - 32px))}
    *{box-sizing:border-box}html{scroll-behavior:smooth}body{margin:0;font-family:"Inter",sans-serif;color:var(--ink);background:linear-gradient(180deg,#f8fafc 0%,#f4f5f7 100%)}img{display:block;width:100%}a{text-decoration:none;color:inherit}button,input,textarea{font:inherit}
    .container{width:var(--container);margin:0 auto}.section{padding:44px 0;position:relative}.section:nth-of-type(even){background:#fff}.eyebrow{display:inline-flex;padding:9px 14px;border-radius:4px;background:var(--brand);font-size:12px;font-weight:800;letter-spacing:.08em;text-transform:uppercase;color:#fff}.section-title{margin:16px 0 12px;font-size:clamp(28px,4vw,44px);line-height:1.06;position:relative;padding-left:18px}.section-title::before{content:"";position:absolute;left:0;top:3px;bottom:3px;width:5px;background:var(--brand)}.section-text{margin:0;color:var(--muted);line-height:1.8;max-width:760px}
    h1,h2,h3,h4,.section-title,.logo strong,.service-card strong,.project-card strong,.quote-card strong,.article-card strong,.type-overlay strong,.process-step strong,.card strong{text-transform:uppercase;letter-spacing:.03em}
    .topbar{background:var(--green);color:rgba(255,255,255,.92);font-size:13px;border-bottom:3px solid var(--brand)}.topbar-inner{display:flex;justify-content:space-between;gap:16px;flex-wrap:wrap;padding:10px 0}
    .header-shell{position:sticky;top:0;z-index:30;padding:0}.header{width:100%;max-width:none;margin:0;display:grid;grid-template-columns:auto 1fr auto;gap:18px;align-items:center;padding:16px max(16px,calc((100vw - 1180px)/2));background:#fff;border-bottom:1px solid var(--line);box-shadow:0 8px 18px rgba(19,32,51,.06)}
    .logo{display:flex;gap:14px;align-items:center}.logo-mark{width:74px;height:74px;display:grid;place-items:center;flex:0 0 74px}.logo-mark svg{width:100%;height:100%;display:block}.logo strong{display:block;font-size:18px}.logo span{display:block;color:var(--muted);font-size:12px;margin-top:2px}
    .nav{display:flex;justify-content:center;gap:8px;flex-wrap:wrap}.nav a{padding:12px 14px;border-radius:0;color:var(--ink);font-weight:700;font-size:13px;text-transform:uppercase;letter-spacing:.05em}.nav a:hover{color:var(--brand-dark)}
    .lang-switch{display:flex;gap:6px;padding:5px;border-radius:4px;background:#eef2f7;border:1px solid var(--line)}.lang-btn{border:0;background:transparent;padding:8px 12px;border-radius:2px;cursor:pointer;font-weight:700;color:var(--muted)}.lang-btn.active{background:#fff;color:var(--green);box-shadow:0 4px 10px rgba(19,32,51,.08)}
    .header-actions{display:flex;gap:10px;flex-wrap:wrap}.btn,.btn-outline{display:inline-flex;align-items:center;justify-content:center;gap:8px;padding:13px 18px;border-radius:4px;border:0;cursor:pointer;font-weight:700;transition:transform .2s ease,box-shadow .2s ease}.btn:hover,.btn-outline:hover{transform:translateY(-1px)}.btn{background:linear-gradient(180deg,var(--brand),var(--brand-dark));color:#fff;box-shadow:0 12px 22px rgba(242,139,32,.18)}.btn-outline{background:#fff;color:var(--green);border:1px solid var(--line)}
    .share-feedback{font-size:13px;color:var(--green);min-height:18px;margin-top:10px}
    .share-cluster{display:flex;gap:10px;flex-wrap:wrap;align-items:center}
    .share-mini{display:inline-flex;align-items:center;justify-content:center;gap:8px;padding:12px 14px;border-radius:4px;border:1px solid var(--line);background:#fff;color:var(--green);font-weight:700;cursor:pointer;text-decoration:none;transition:transform .2s ease,box-shadow .2s ease}
    .share-mini:hover{transform:translateY(-1px);box-shadow:0 12px 24px rgba(20,34,29,.08)}
    .share-mini svg{width:18px;height:18px;display:block}
    .share-mini.zalo{color:#0068ff}
    .share-mini.facebook{color:#1877f2}
    .share-mini.messenger{color:#0084ff}
    .hero-grid{display:grid;grid-template-columns:minmax(0,1.32fr) 360px;gap:22px;align-items:stretch}.hero-stage{position:relative;min-height:610px;overflow:hidden;border-radius:0;box-shadow:var(--shadow);border:1px solid var(--line)}.slide{position:absolute;inset:0;opacity:0;transition:opacity .55s ease}.slide.active{opacity:1}.slide::after{content:"";position:absolute;inset:0;background:linear-gradient(90deg,rgba(16,29,47,.86) 0%,rgba(16,29,47,.42) 54%,rgba(16,29,47,.16) 100%)}.slide img{height:100%;object-fit:cover}.slide-content{position:absolute;left:38px;bottom:38px;width:min(580px,calc(100% - 76px));color:#fff;z-index:1}.slide-content p{color:rgba(255,255,255,.88);line-height:1.75;max-width:520px}.dots{position:absolute;top:0;right:0;left:auto;display:flex;gap:0;z-index:2}.dot{width:54px;height:54px;border-radius:0;border:0;background:rgba(255,255,255,.18);cursor:pointer}.dot.active{width:54px;background:var(--brand)}
    .panel,.card,.service-card,.project-card,.quote-card,.article-card,.contact-card{background:var(--surface);border:1px solid var(--line);border-radius:0;box-shadow:var(--shadow)}.panel,.contact-card,.card,.quote-card,.article-card{padding:26px}.hero-side{display:grid;gap:18px}.meta-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}.meta-box{padding:18px;border-radius:0;background:var(--surface-soft);border:1px solid var(--line)}.meta-box strong{display:block;font-size:24px;margin-top:10px}.meta-box span{font-size:13px;color:var(--muted);text-transform:uppercase;font-weight:700;letter-spacing:.06em}
    .highlights{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:18px}.card strong,.service-card strong,.project-card strong,.quote-card strong,.article-card strong{display:block}
    .services-grid,.project-grid,.quote-grid,.article-grid{display:grid;gap:20px}.services-grid{grid-template-columns:repeat(4,minmax(0,1fr))}.service-card,.project-card{overflow:hidden;position:relative}.service-card::before,.project-card::before,.quote-card::before,.article-card::before,.card::before,.contact-card::before{content:"";position:absolute;inset:0 auto auto 0;width:100%;height:4px;background:var(--brand)}.service-card img,.project-card img{height:240px;object-fit:cover}.service-body,.project-body{padding:20px}
    .types-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:20px}.type-card{position:relative;overflow:hidden;min-height:300px;border-radius:0;box-shadow:var(--shadow);border:1px solid var(--line)}.type-card img{height:100%;object-fit:cover;transition:transform .45s ease}.type-card:hover img{transform:scale(1.05)}.type-overlay{position:absolute;inset:auto 0 0 0;padding:22px;color:#fff;background:linear-gradient(180deg,rgba(0,0,0,0) 0%,rgba(16,29,47,.9) 100%)}.type-overlay strong{font-size:22px}.type-overlay span{display:block;color:rgba(255,255,255,.82);margin-top:6px;line-height:1.5}
    .process{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:18px}.process-step{padding:24px;border-radius:0;background:#fff;border:1px solid var(--line);box-shadow:0 12px 28px rgba(19,32,51,.06)}.process-step em{display:inline-flex;width:48px;height:48px;border-radius:0;background:#eef2f7;align-items:center;justify-content:center;color:var(--brand-dark);font-style:normal;font-weight:800;border-left:4px solid var(--brand)}
    .showcase{display:grid;grid-template-columns:1.02fr .98fr;gap:22px}.gallery{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px;margin-top:18px}.gallery img{height:190px;object-fit:cover;border-radius:0}
    .quote-grid,.article-grid{grid-template-columns:repeat(3,minmax(0,1fr))}.article-card img{height:210px;object-fit:cover;border-radius:0;margin-bottom:16px}.article-date{display:block;margin:0 0 8px;color:var(--brand-dark);font-size:12px;font-weight:800;letter-spacing:.08em;text-transform:uppercase}
    .contact-wrap{display:grid;grid-template-columns:1fr 420px;gap:22px}.contact-actions{display:flex;gap:12px;flex-wrap:wrap;margin-top:18px}.contact-list{display:grid;gap:10px;color:var(--muted);line-height:1.8}.contact-form{display:grid;gap:12px}.contact-form input,.contact-form textarea{width:100%;padding:13px 14px;border-radius:0;border:1px solid var(--line);background:#fff;outline:0}.contact-form textarea{min-height:110px;resize:vertical}.status{min-height:20px;font-size:14px;color:var(--green)}.status.error{color:#b14b2f}
    .footer{margin-top:18px;padding:38px 0 42px;background:#13243b;color:rgba(255,255,255,.82);border-top:4px solid var(--brand)}.footer-grid{display:grid;grid-template-columns:1.3fr 1fr 1fr;gap:24px}.footer h4{margin:0 0 12px;color:#fff}.footer p,.footer li{line-height:1.75;color:rgba(255,255,255,.72)}.footer ul{list-style:none;padding:0;margin:0;display:grid;gap:8px}
    .floating{position:fixed;right:18px;bottom:18px;z-index:20;display:grid;gap:10px}.floating a{padding:12px 16px;border-radius:4px;background:var(--brand);color:#fff;box-shadow:0 12px 26px rgba(0,0,0,.16)}
    @media (max-width:1120px){.header,.hero-grid,.contact-wrap,.showcase,.footer-grid{grid-template-columns:1fr}.highlights,.services-grid,.types-grid,.process,.quote-grid,.article-grid{grid-template-columns:repeat(2,minmax(0,1fr))}}
    @media (max-width:720px){.topbar{display:none}.header-shell{padding-top:8px}.header{padding:14px;border-radius:20px}.nav{justify-content:flex-start}.section{padding:26px 0}.highlights,.services-grid,.types-grid,.process,.quote-grid,.article-grid,.meta-grid,.gallery{grid-template-columns:1fr}.hero-stage{min-height:520px}.slide-content{left:20px;bottom:20px;width:calc(100% - 40px)}.header-actions{display:none}}
  </style>
</head>
<body>
  <div class="topbar">
    <div class="container topbar-inner">
      <div><span id="companyHours"></span> • <span id="companyAddress"></span></div>
      <div>Hotline/Zalo: <span id="companyPhoneTop"></span></div>
    </div>
  </div>

  <div class="header-shell">
    <header class="header">
      <a class="logo" href="#top">
        <div class="logo-mark" aria-hidden="true">
          <svg viewBox="0 0 160 160" xmlns="http://www.w3.org/2000/svg" role="img">
            <rect x="18" y="26" width="58" height="92" fill="#84e58b"/>
            <rect x="34" y="42" width="58" height="76" fill="#0d7d1b"/>
            <rect x="62" y="52" width="20" height="66" fill="#ffffff"/>
            <rect x="88" y="46" width="56" height="16" fill="#2dbf43"/>
            <rect x="88" y="64" width="16" height="54" fill="#0d7d1b"/>
            <rect x="108" y="64" width="16" height="54" fill="#1e962d"/>
            <rect x="128" y="64" width="16" height="54" fill="#49d35b"/>
            <polygon points="28,34 92,10 92,26 28,50" fill="#0b6517"/>
            <polygon points="72,48 148,76 148,118 72,90" fill="#158f24"/>
            <rect x="38" y="80" width="8" height="8" fill="#ffffff"/>
            <rect x="50" y="80" width="8" height="8" fill="#ffffff"/>
            <rect x="38" y="92" width="8" height="8" fill="#ffffff"/>
            <rect x="50" y="92" width="8" height="8" fill="#ffffff"/>
            <rect x="38" y="104" width="8" height="8" fill="#ffffff"/>
            <rect x="50" y="104" width="8" height="8" fill="#ffffff"/>
            <rect x="20" y="118" width="126" height="6" fill="#0d7d1b"/>
          </svg>
        </div>
        <div>
          <strong id="companyName"></strong>
          <span id="companyTagline"></span>
        </div>
      </a>
      <nav class="nav">
        <a id="navServices" href="#services">Dich vu</a>
        <a id="navProjects" href="#projects">Du an</a>
        <a id="navProcess" href="#process">Quy trinh</a>
        <a id="navArticles" href="#articles">Cam nang</a>
        <a id="navContact" href="#contact">Lien he</a>
      </nav>
      <div class="header-actions">
        <div class="lang-switch">
          <button class="lang-btn active" id="langVi" type="button">VI</button>
          <button class="lang-btn" id="langEn" type="button">EN</button>
        </div>
        <a class="btn-outline" id="zaloLinkHeader" href="#" target="_blank" rel="noreferrer">Zalo</a>
        <a class="btn" id="phoneLinkHeader" href="#">Goi ngay</a>
      </div>
    </header>
  </div>

  <main>
    <section class="section" id="top">
      <div class="container hero-grid">
        <div class="hero-stage">
          <div class="dots" id="heroDots"></div>
          <div id="heroSlides"></div>
        </div>
        <aside class="hero-side">
          <div class="panel">
            <span class="eyebrow" id="heroPanelEyebrow">Don vi thiet ke va thi cong</span>
            <h2 id="heroPanelTitle" style="margin:14px 0 12px;font-size:28px;">Giai phap tron goi cho nha pho, noi that va cai tao cong trinh</h2>
            <p class="section-text" id="websiteNote"></p>
            <div class="meta-grid">
              <div class="meta-box"><span id="metaLabel1">Linh vuc</span><strong>4+</strong></div>
              <div class="meta-box"><span id="metaLabel2">Hang muc</span><strong id="metaValue2">Tron goi</strong></div>
              <div class="meta-box"><span id="metaLabel3">Phong cach</span><strong id="metaValue3">Hien dai</strong></div>
              <div class="meta-box"><span id="metaLabel4">Lien he</span><strong id="companyPhoneBox"></strong></div>
            </div>
          </div>
          <div class="contact-card">
            <span class="eyebrow" id="consultEyebrow">Yeu cau tu van</span>
            <h3 id="consultTitle" style="margin:14px 0 12px;">De lai thong tin, TH-Ecohome se lien he lai</h3>
            <form class="contact-form" id="consultForm">
              <input id="consultName" type="text" placeholder="Ho va ten" required />
              <input id="consultPhone" type="tel" placeholder="So dien thoai" required />
              <textarea id="consultMessage" placeholder="Ban dang quan tam thiet ke, thi cong hay cai tao hang muc nao?"></textarea>
              <button class="btn" id="consultSubmit" type="submit">Gui yeu cau</button>
              <div class="status" id="consultStatus"></div>
            </form>
          </div>
        </aside>
      </div>
    </section>

    <section class="section">
      <div class="container highlights" id="highlightGrid"></div>
    </section>

    <section class="section" id="services">
      <div class="container">
        <span class="eyebrow" id="servicesEyebrow">Dich vu chinh</span>
        <h2 class="section-title" id="servicesTitle">TH-Ecohome tap trung vao cac hang muc sat voi nhu cau thuc te</h2>
        <p class="section-text" id="servicesText">Huong toi cach lam viec ro rang, hien dai, de hieu va de phoi hop cho chu nha, gia dinh tre va cong trinh quy mo vua.</p>
        <div class="services-grid" id="servicesGrid" style="margin-top:20px;"></div>
      </div>
    </section>

    <section class="section">
      <div class="container">
        <span class="eyebrow" id="typesEyebrow">Loai cong trinh</span>
        <h2 class="section-title" id="typesTitle">Khong gian TH-Ecohome co the dong hanh</h2>
        <p class="section-text" id="typesText">Tu nha pho dan dung den noi that can ho, bep, phong tam va cac hang muc cai tao can doi moi nhanh gon.</p>
        <div class="types-grid" id="typeGrid" style="margin-top:20px;"></div>
      </div>
    </section>

    <section class="section" id="projects">
      <div class="container showcase">
        <div class="card">
          <span class="eyebrow" id="projectsEyebrow">Du an tieu bieu</span>
          <h2 id="projectsTitle" style="margin:14px 0 10px;font-size:34px;">Huong den khong gian song gon, sang va de su dung lau dai</h2>
          <p class="section-text" id="projectsText">Website nay duoc doi noi dung theo dung nganh nghe cua TH-Ecohome: nha pho dan dung, noi that, cai tao va thi cong hoan thien.</p>
          <div class="gallery">
            <img src="https://images.unsplash.com/photo-1600047509807-ba8f99d2cdde?auto=format&fit=crop&w=900&q=80" alt="nha pho" />
            <img src="https://images.unsplash.com/photo-1494526585095-c41746248156?auto=format&fit=crop&w=900&q=80" alt="noi that" />
            <img src="https://images.unsplash.com/photo-1556911220-bff31c812dba?auto=format&fit=crop&w=900&q=80" alt="bep" />
            <img src="https://images.unsplash.com/photo-1484154218962-a197022b5858?auto=format&fit=crop&w=900&q=80" alt="cai tao" />
          </div>
        </div>
        <div class="project-grid" id="projectGrid"></div>
      </div>
    </section>

    <section class="section" id="process">
      <div class="container">
        <span class="eyebrow" id="processEyebrow">Quy trinh</span>
        <h2 class="section-title" id="processTitle">Lam viec ro rang tu dau de de kiem soat chat luong va tien do</h2>
        <div class="process" id="processGrid" style="margin-top:20px;"></div>
      </div>
    </section>

    <section class="section">
      <div class="container">
        <span class="eyebrow" id="quotesEyebrow">Danh gia</span>
        <h2 class="section-title" id="quotesTitle">Cam nhan tu khach hang</h2>
        <div class="quote-grid" id="quoteGrid" style="margin-top:20px;"></div>
      </div>
    </section>

    <section class="section" id="articles">
      <div class="container">
        <span class="eyebrow" id="articlesEyebrow">Cam nang</span>
        <h2 class="section-title" id="articlesTitle">Noi dung goi y de xay dung niem tin va chia se kinh nghiem</h2>
        <div class="article-grid" id="articleGrid" style="margin-top:20px;"></div>
      </div>
    </section>

    <section class="section" id="contact">
      <div class="container contact-wrap">
        <div class="card">
          <span class="eyebrow" id="contactEyebrow">Lien he day du</span>
          <h2 id="contactTitle" style="margin:14px 0 12px;font-size:34px;">Cong ty TNHH Thuong Mai va Dich vu TH-Ecohome</h2>
          <p class="section-text" id="contactText">Neu ban can tu van thiet ke, thi cong nha pho, noi that hay cai tao cong trinh, co the lien he truc tiep qua dien thoai hoac Zalo.</p>
          <div class="contact-list" style="margin-top:18px;">
            <div><strong id="contactAddressLabel">Tru so:</strong> <span id="contactAddress"></span></div>
            <div><strong id="contactPhoneLabel">SDT lien he:</strong> <span id="contactPhone"></span></div>
            <div><strong>Zalo:</strong> <span id="contactZalo"></span></div>
            <div><strong id="contactHoursLabel">Gio tu van:</strong> <span id="contactHours"></span></div>
            <div><strong id="contactFieldLabel">Linh vuc:</strong> <span id="contactFieldText">Thiet ke, thi cong cong trinh nha pho dan dung, noi that, cai tao cong trinh.</span></div>
          </div>
          <div class="contact-actions">
            <a class="btn" id="phoneLinkContact" href="#">Goi 0367954406</a>
            <a class="btn-outline" id="zaloLinkContact" href="#" target="_blank" rel="noreferrer">Nhan Zalo</a>
          </div>
        </div>
        <div class="contact-card">
          <span class="eyebrow" id="contactFormEyebrow">Hen lich nhanh</span>
          <h3 id="contactFormTitle" style="margin:14px 0 12px;">Gui thong tin de duoc lien he lai</h3>
          <form class="contact-form" id="contactForm">
            <input id="contactName" type="text" placeholder="Ho va ten" required />
            <input id="contactPhoneInput" type="tel" placeholder="So dien thoai" required />
            <textarea id="contactMessage" placeholder="Mo ta ngan nhu cau cong trinh cua ban"></textarea>
            <button class="btn" id="contactSubmit" type="submit">Gui thong tin</button>
            <div class="status" id="contactStatus"></div>
          </form>
        </div>
      </div>
    </section>
  </main>

  <footer class="footer">
    <div class="container footer-grid">
      <div>
        <div style="display:flex;align-items:center;gap:12px;margin-bottom:10px;">
          <div class="logo-mark" style="width:62px;height:62px;flex-basis:62px;">
            <svg viewBox="0 0 160 160" xmlns="http://www.w3.org/2000/svg" role="img">
              <rect x="18" y="26" width="58" height="92" fill="#84e58b"/>
              <rect x="34" y="42" width="58" height="76" fill="#0d7d1b"/>
              <rect x="62" y="52" width="20" height="66" fill="#ffffff"/>
              <rect x="88" y="46" width="56" height="16" fill="#2dbf43"/>
              <rect x="88" y="64" width="16" height="54" fill="#0d7d1b"/>
              <rect x="108" y="64" width="16" height="54" fill="#1e962d"/>
              <rect x="128" y="64" width="16" height="54" fill="#49d35b"/>
              <polygon points="28,34 92,10 92,26 28,50" fill="#0b6517"/>
              <polygon points="72,48 148,76 148,118 72,90" fill="#158f24"/>
              <rect x="38" y="80" width="8" height="8" fill="#ffffff"/>
              <rect x="50" y="80" width="8" height="8" fill="#ffffff"/>
              <rect x="38" y="92" width="8" height="8" fill="#ffffff"/>
              <rect x="50" y="92" width="8" height="8" fill="#ffffff"/>
              <rect x="38" y="104" width="8" height="8" fill="#ffffff"/>
              <rect x="50" y="104" width="8" height="8" fill="#ffffff"/>
              <rect x="20" y="118" width="126" height="6" fill="#0d7d1b"/>
            </svg>
          </div>
          <h4 id="footerName" style="margin:0;"></h4>
        </div>
        <p id="footerTagline"></p>
      </div>
      <div>
        <h4 id="footerContactTitle">Thong tin lien he</h4>
        <ul>
          <li id="footerAddress"></li>
          <li id="footerPhone"></li>
          <li id="footerZalo"></li>
          <li id="footerHours"></li>
        </ul>
      </div>
      <div>
        <h4 id="footerFieldTitle">Linh vuc chuyen mon</h4>
        <ul>
          <li id="footerField1">Thiet ke nha pho dan dung</li>
          <li id="footerField2">Thi cong cong trinh va hoan thien</li>
          <li id="footerField3">Thiet ke va thi cong noi that</li>
          <li id="footerField4">Cai tao cong trinh va khong gian song</li>
        </ul>
      </div>
    </div>
  </footer>

  <div class="floating">
    <a id="floatingTop" href="#top">Len dau trang</a>
    <a id="zaloFloating" href="#" target="_blank" rel="noreferrer">Zalo</a>
  </div>

  <script>const storeData=__STORE_DATA__;</script>
  <script>
    const pageText={
      vi:{
        lang:"vi", title:"TH-Ecohome | Thiết kế, thi công, nội thất và cải tạo công trình",
        topbarLabel:"Hotline/Zalo:",
        navServices:"Dịch vụ", navProjects:"Dự án", navProcess:"Quy trình", navArticles:"Cẩm nang", navContact:"Liên hệ",
        callNow:"Gọi ngay", heroPanelEyebrow:"Đơn vị thiết kế và thi công", heroPanelTitle:"Giải pháp trọn gói cho nhà phố, nội thất và cải tạo công trình",
        metaLabel1:"Lĩnh vực", metaLabel2:"Hạng mục", metaValue2:"Trọn gói", metaLabel3:"Phong cách", metaValue3:"Hiện đại", metaLabel4:"Liên hệ",
        consultEyebrow:"Yêu cầu tư vấn", consultTitle:"Để lại thông tin, TH-Ecohome sẽ liên hệ lại", consultSubmit:"Gửi yêu cầu",
        consultName:"Họ và tên", consultPhone:"Số điện thoại", consultMessage:"Bạn đang quan tâm thiết kế, thi công hay cải tạo hạng mục nào?",
        servicesEyebrow:"Dịch vụ chính", servicesTitle:"TH-Ecohome tập trung vào các hạng mục sát với nhu cầu thực tế", servicesText:"Hướng tới cách làm việc rõ ràng, hiện đại, dễ hiểu và dễ phối hợp cho chủ nhà, gia đình trẻ và công trình quy mô vừa.",
        typesEyebrow:"Loại công trình", typesTitle:"Không gian TH-Ecohome có thể đồng hành", typesText:"Từ nhà phố dân dụng đến nội thất căn hộ, bếp, phòng tắm và các hạng mục cải tạo cần đổi mới nhanh gọn.",
        projectsEyebrow:"Dự án tiêu biểu", projectsTitle:"Hướng đến không gian sống gọn, sáng và dễ sử dụng lâu dài", projectsText:"Website này được đổi nội dung theo đúng ngành nghề của TH-Ecohome: nhà phố dân dụng, nội thất, cải tạo và thi công hoàn thiện.",
        processEyebrow:"Quy trình", processTitle:"Làm việc rõ ràng từ đầu để dễ kiểm soát chất lượng và tiến độ",
        quotesEyebrow:"Đánh giá", quotesTitle:"Cảm nhận từ khách hàng",
        articlesEyebrow:"Cẩm nang", articlesTitle:"Nội dung gợi ý để xây dựng niềm tin và chia sẻ kinh nghiệm",
        contactEyebrow:"Liên hệ đầy đủ", contactTitle:"Công ty TNHH Thương Mại và Dịch vụ TH-Ecohome", contactText:"Nếu bạn cần tư vấn thiết kế, thi công nhà phố, nội thất hay cải tạo công trình, có thể liên hệ trực tiếp qua điện thoại hoặc Zalo.",
        contactAddressLabel:"Trụ sở:", contactPhoneLabel:"SĐT liên hệ:", contactHoursLabel:"Giờ tư vấn:", contactFieldLabel:"Lĩnh vực:", contactFieldText:"Thiết kế, thi công công trình nhà phố dân dụng, nội thất, cải tạo công trình.",
        phoneContact:"Gọi 0367954406", zaloContact:"Nhắn Zalo",
        contactFormEyebrow:"Hẹn lịch nhanh", contactFormTitle:"Gửi thông tin để được liên hệ lại", contactSubmit:"Gửi thông tin", contactName:"Họ và tên", contactPhoneInput:"Số điện thoại", contactMessage:"Mô tả ngắn nhu cầu công trình của bạn",
        footerContactTitle:"Thông tin liên hệ", footerFieldTitle:"Lĩnh vực chuyên môn",
        footerField1:"Thiết kế nhà phố dân dụng", footerField2:"Thi công công trình và hoàn thiện", footerField3:"Thiết kế và thi công nội thất", footerField4:"Cải tạo công trình và không gian sống",
        floatingTop:"Lên đầu trang", heroPrimary:"Nhận tư vấn", heroSecondary:"Xem dịch vụ",
        consultOk:(name,phone)=>`Đã ghi nhận yêu cầu của ${name}. TH-Ecohome sẽ liên hệ qua ${phone}.`,
        consultErr:"Vui lòng nhập họ tên và số điện thoại.",
        contactOk:(name)=>`Thông tin của ${name} đã được ghi nhận. Bạn có thể gọi trực tiếp ${storeData.company.phone} hoặc nhắn Zalo để trao đổi nhanh hơn.`,
        contactErr:"Vui lòng nhập đầy đủ họ tên và số điện thoại."
      },
      en:{
        lang:"en", title:"TH-Ecohome | Design, construction, interiors and renovation",
        topbarLabel:"Hotline/Zalo:",
        navServices:"Services", navProjects:"Projects", navProcess:"Process", navArticles:"Insights", navContact:"Contact",
        callNow:"Call now", heroPanelEyebrow:"Design and build studio", heroPanelTitle:"Turnkey solutions for townhouses, interiors and renovation projects",
        metaLabel1:"Focus", metaLabel2:"Scope", metaValue2:"Turnkey", metaLabel3:"Style", metaValue3:"Modern", metaLabel4:"Contact",
        consultEyebrow:"Request consultation", consultTitle:"Leave your details and TH-Ecohome will contact you", consultSubmit:"Send request",
        consultName:"Full name", consultPhone:"Phone number", consultMessage:"Which design, construction or renovation package are you interested in?",
        servicesEyebrow:"Core services", servicesTitle:"TH-Ecohome focuses on practical services for real living needs", servicesText:"A clear, modern workflow built for homeowners, young families and medium-scale residential projects.",
        typesEyebrow:"Project types", typesTitle:"Spaces TH-Ecohome can support", typesText:"From townhouses to apartment interiors, kitchens, bathrooms and fast renovation packages.",
        projectsEyebrow:"Featured projects", projectsTitle:"Aiming for clean, bright and lasting living spaces", projectsText:"This website is tailored to TH-Ecohome's actual field: residential townhouses, interiors, renovation and finishing works.",
        processEyebrow:"Workflow", processTitle:"A clear process from day one helps control quality and timeline",
        quotesEyebrow:"Testimonials", quotesTitle:"What clients say",
        articlesEyebrow:"Insights", articlesTitle:"Suggested content to build trust and share practical experience",
        contactEyebrow:"Full contact", contactTitle:"TH-Ecohome Trading and Services Co., Ltd.", contactText:"If you need advice on townhouse design, construction, interiors or renovation, you can contact us directly by phone or Zalo.",
        contactAddressLabel:"Address:", contactPhoneLabel:"Phone:", contactHoursLabel:"Consulting hours:", contactFieldLabel:"Field:", contactFieldText:"Townhouse design, construction, interior design-build and renovation works.",
        phoneContact:"Call 0367954406", zaloContact:"Chat on Zalo",
        contactFormEyebrow:"Quick booking", contactFormTitle:"Send your information for a callback", contactSubmit:"Send information", contactName:"Full name", contactPhoneInput:"Phone number", contactMessage:"Briefly describe your project needs",
        footerContactTitle:"Contact information", footerFieldTitle:"Specialized fields",
        footerField1:"Residential townhouse design", footerField2:"Construction and finishing works", footerField3:"Interior design and build", footerField4:"Renovation and space upgrades",
        floatingTop:"Back to top", heroPrimary:"Get consultation", heroSecondary:"View services",
        consultOk:(name,phone)=>`Your request from ${name} has been received. TH-Ecohome will contact you at ${phone}.`,
        consultErr:"Please enter your full name and phone number.",
        contactOk:(name)=>`${name}'s information has been received. You can also call ${storeData.company.phone} or message Zalo for faster support.`,
        contactErr:"Please enter your full name and phone number."
      }
    };
    const localizedContent={
      en:{
        company:{tagline:"Design, construction, interiors and renovation for modern living", hours:"Consultation daily: 08:00 - 20:00", website_note:"A team specializing in residential design-build, interiors and practical renovation for modern homes."},
        hero_slides:[
          {eyebrow:"TH-Ecohome", title:"Minimal and youthful townhouse design-build with practical layouts", description:"From concept, spatial planning and 3D visuals to completed construction for residential projects."},
          {eyebrow:"Interiors and finishing", title:"Clean, modern spaces with optimized cost for homes and apartments", description:"Focused on durable materials, modern tones and interior solutions that fit real daily use."},
          {eyebrow:"Renovation", title:"Refresh old homes, facades, kitchens, bathrooms and business spaces", description:"Neat execution, transparent work items and priority on schedule so spaces can return to use quickly."}
        ],
        highlights:[
          {title:"Townhouse design", text:"Unified solutions for layouts, facades and interiors."},
          {title:"Turnkey construction", text:"Close timeline control, work-item management and finishing quality."},
          {title:"Modern interiors", text:"Minimal, youthful and easy to maintain."},
          {title:"Flexible renovation", text:"Functional upgrades for houses, apartments and small projects."}
        ],
        services:[
          {name:"Townhouse architectural design", summary:"Functional planning, facade concepts, visuals and coordinated material direction."},
          {name:"Residential construction", summary:"Structure, finishing and coordinated work items delivered on plan."},
          {name:"Interior design and build", summary:"Modern living room, kitchen, bedroom and bathroom solutions."},
          {name:"Renovation works", summary:"Repair, upgrade, functional adjustment and space renewal."}
        ],
        project_types:[
          {name:"Residential townhouses", style:"Clean facades, balanced proportions and natural light"},
          {name:"Apartment interiors", style:"Minimal, youthful and space-efficient"},
          {name:"Kitchens and bathrooms", style:"Function-first, durable and easy to clean"},
          {name:"Facade renovation", style:"Fast and effective visual refresh"}
        ],
        projects:[
          {name:"Three-storey modern townhouse", tag:"Design + Build", text:"Optimized ventilation, natural light and shared family living areas."},
          {name:"Minimal two-bedroom apartment", tag:"Interior", text:"Reorganized living room-kitchen flow with light wood tones and neutral finishes."},
          {name:"Kitchen and bathroom upgrade", tag:"Renovation", text:"Upgraded materials, lighting, fixtures and neat storage solutions for long-used homes."}
        ],
        process:[
          {title:"Requirement intake", text:"Discuss goals, budget, style and current site conditions."},
          {title:"Concept planning", text:"Propose design direction, layout and suitable materials."},
          {title:"Construction rollout", text:"Finalize work items, schedule and supervise each phase."},
          {title:"Handover and support", text:"Inspection, handover and ongoing support after completion."}
        ],
        articles:[
          {title:"3 notes when designing townhouses for young families", excerpt:"Balance function, ventilation, daylight and future flexibility."},
          {title:"How to renovate a kitchen beautifully while keeping it neat", excerpt:"Focus on circulation, easy-clean surfaces and practical storage."},
          {title:"Minimal interior style for modern homes", excerpt:"Use restrained colors, durable materials and lighting for a spacious feel."}
        ],
        testimonials:[
          {role:"Homeowner", quote:"TH-Ecohome worked clearly and was very easy to coordinate with during our renovation."},
          {role:"Interior client", quote:"The style is youthful and modern, and the interior solutions fit our daily life well."},
          {role:"Townhouse client", quote:"From consultation to construction, everything was tidy, clear and easy to follow."}
        ]
      }
    };
    const normalizedVietnameseText={
      title:"TH-Ecohome | Thiết kế, thi công, nội thất và cải tạo công trình",
      topbarLabel:"Hotline/Zalo:",
      navServices:"Dịch vụ",
      navProjects:"Dự án",
      navProcess:"Quy trình",
      navArticles:"Cẩm nang",
      navContact:"Liên hệ",
      callNow:"Gọi ngay",
      heroPanelEyebrow:"Đơn vị thiết kế và thi công",
      heroPanelTitle:"Giải pháp trọn gói cho nhà phố, nội thất và cải tạo công trình",
      metaLabel1:"Lĩnh vực",
      metaLabel2:"Hạng mục",
      metaValue2:"Trọn gói",
      metaLabel3:"Phong cách",
      metaValue3:"Hiện đại",
      metaLabel4:"Liên hệ",
      consultEyebrow:"Yêu cầu tư vấn",
      consultTitle:"Để lại thông tin, TH-Ecohome sẽ liên hệ lại",
      consultSubmit:"Gửi yêu cầu",
      consultName:"Họ và tên",
      consultPhone:"Số điện thoại",
      consultMessage:"Bạn đang quan tâm thiết kế, thi công hay cải tạo hạng mục nào?",
      servicesEyebrow:"Dịch vụ chính",
      servicesTitle:"TH-Ecohome tập trung vào các hạng mục sát với nhu cầu thực tế",
      servicesText:"Hướng tới cách làm việc rõ ràng, hiện đại, dễ hiểu và dễ phối hợp cho chủ nhà, gia đình trẻ và công trình quy mô vừa.",
      typesEyebrow:"Loại công trình",
      typesTitle:"Không gian TH-Ecohome có thể đồng hành",
      typesText:"Từ nhà phố dân dụng đến nội thất căn hộ, bếp, phòng tắm và các hạng mục cải tạo cần đổi mới nhanh gọn.",
      projectsEyebrow:"Dự án tiêu biểu",
      projectsTitle:"Hướng đến không gian sống gọn, sáng và dễ sử dụng lâu dài",
      projectsText:"Website này được đổi nội dung theo đúng ngành nghề của TH-Ecohome: nhà phố dân dụng, nội thất, cải tạo và thi công hoàn thiện.",
      processEyebrow:"Quy trình",
      processTitle:"Làm việc rõ ràng từ đầu để dễ kiểm soát chất lượng và tiến độ",
      quotesEyebrow:"Đánh giá",
      quotesTitle:"Cảm nhận từ khách hàng",
      articlesEyebrow:"Cẩm nang",
      articlesTitle:"Nội dung gợi ý để xây dựng niềm tin và chia sẻ kinh nghiệm",
      contactEyebrow:"Liên hệ đầy đủ",
      contactTitle:"Công ty TNHH Thương Mại và Dịch vụ TH-Ecohome",
      contactText:"Nếu bạn cần tư vấn thiết kế, thi công nhà phố, nội thất hay cải tạo công trình, có thể liên hệ trực tiếp qua điện thoại hoặc Zalo.",
      contactAddressLabel:"Trụ sở:",
      contactPhoneLabel:"SĐT liên hệ:",
      contactHoursLabel:"Giờ tư vấn:",
      contactFieldLabel:"Lĩnh vực:",
      contactFieldText:"Thiết kế, thi công công trình nhà phố dân dụng, nội thất, cải tạo công trình.",
      phoneContact:"Gọi 0367954406",
      zaloContact:"Nhắn Zalo",
      contactFormEyebrow:"Hẹn lịch nhanh",
      contactFormTitle:"Gửi thông tin để được liên hệ lại",
      contactSubmit:"Gửi thông tin",
      contactName:"Họ và tên",
      contactPhoneInput:"Số điện thoại",
      contactMessage:"Mô tả ngắn nhu cầu công trình của bạn",
      footerContactTitle:"Thông tin liên hệ",
      footerFieldTitle:"Lĩnh vực chuyên môn",
      footerField1:"Thiết kế nhà phố dân dụng",
      footerField2:"Thi công công trình và hoàn thiện",
      footerField3:"Thiết kế và thi công nội thất",
      footerField4:"Cải tạo công trình và không gian sống",
      floatingTop:"Lên đầu trang",
      heroPrimary:"Nhận tư vấn",
      heroSecondary:"Xem dịch vụ",
      consultErr:"Vui lòng nhập họ tên và số điện thoại.",
      contactErr:"Vui lòng nhập đầy đủ họ tên và số điện thoại."
    };
    Object.assign(pageText.vi, normalizedVietnameseText);
    pageText.vi.consultOk=(name,phone)=>`Đã ghi nhận yêu cầu của ${name}. TH-Ecohome sẽ liên hệ qua ${phone}.`;
    pageText.vi.contactOk=(name)=>`Thông tin của ${name} đã được ghi nhận. Bạn có thể gọi trực tiếp ${storeData.company.phone} hoặc nhắn Zalo để trao đổi nhanh hơn.`;
    const phoneHref=`tel:${storeData.company.phone}`;
    const zaloHref=`https://zalo.me/${storeData.company.zalo}`;
    let slideIndex=0;
    let slideTimer=null;
    let currentLang=localStorage.getItem("thecohome_lang") || "vi";

    function getContent(){
      if(currentLang==="vi"){return storeData;}
      const en=localizedContent.en;
      return {
        ...storeData,
        company:{...storeData.company,...en.company},
        hero_slides:storeData.hero_slides.map((item,index)=>({...item,...en.hero_slides[index]})),
        highlights:storeData.highlights.map((item,index)=>({...item,...en.highlights[index]})),
        services:storeData.services.map((item,index)=>({...item,...en.services[index]})),
        project_types:storeData.project_types.map((item,index)=>({...item,...en.project_types[index]})),
        projects:storeData.projects.map((item,index)=>({...item,...en.projects[index]})),
        process:storeData.process.map((item,index)=>({...item,...en.process[index]})),
        articles:storeData.articles.map((item,index)=>({...item,...en.articles[index]})),
        testimonials:storeData.testimonials.map((item,index)=>({...item,...en.testimonials[index]})),
      };
    }

    function applyStaticText(){
      const t=pageText[currentLang];
      document.documentElement.lang=t.lang;
      document.title=t.title;
      document.getElementById("navServices").textContent=t.navServices;
      document.getElementById("navProjects").textContent=t.navProjects;
      document.getElementById("navProcess").textContent=t.navProcess;
      document.getElementById("navArticles").textContent=t.navArticles;
      document.getElementById("navContact").textContent=t.navContact;
      const shareButton=document.getElementById("shareButton");
      const copyLinkButton=document.getElementById("copyLinkButton");
      if(shareButton){shareButton.setAttribute("aria-label",currentLang==="vi"?"Chia sẻ link website":"Share website link");}
      if(copyLinkButton){copyLinkButton.setAttribute("aria-label",currentLang==="vi"?"Copy link website":"Copy website link");}
      document.getElementById("phoneLinkHeader").textContent=t.callNow;
      document.getElementById("heroPanelEyebrow").textContent=t.heroPanelEyebrow;
      document.getElementById("heroPanelTitle").textContent=t.heroPanelTitle;
      document.getElementById("metaLabel1").textContent=t.metaLabel1;
      document.getElementById("metaLabel2").textContent=t.metaLabel2;
      document.getElementById("metaValue2").textContent=t.metaValue2;
      document.getElementById("metaLabel3").textContent=t.metaLabel3;
      document.getElementById("metaValue3").textContent=t.metaValue3;
      document.getElementById("metaLabel4").textContent=t.metaLabel4;
      document.getElementById("consultEyebrow").textContent=t.consultEyebrow;
      document.getElementById("consultTitle").textContent=t.consultTitle;
      document.getElementById("consultName").placeholder=t.consultName;
      document.getElementById("consultPhone").placeholder=t.consultPhone;
      document.getElementById("consultMessage").placeholder=t.consultMessage;
      document.getElementById("consultSubmit").textContent=t.consultSubmit;
      document.getElementById("servicesEyebrow").textContent=t.servicesEyebrow;
      document.getElementById("servicesTitle").textContent=t.servicesTitle;
      document.getElementById("servicesText").textContent=t.servicesText;
      document.getElementById("typesEyebrow").textContent=t.typesEyebrow;
      document.getElementById("typesTitle").textContent=t.typesTitle;
      document.getElementById("typesText").textContent=t.typesText;
      document.getElementById("projectsEyebrow").textContent=t.projectsEyebrow;
      document.getElementById("projectsTitle").textContent=t.projectsTitle;
      document.getElementById("projectsText").textContent=t.projectsText;
      document.getElementById("processEyebrow").textContent=t.processEyebrow;
      document.getElementById("processTitle").textContent=t.processTitle;
      document.getElementById("quotesEyebrow").textContent=t.quotesEyebrow;
      document.getElementById("quotesTitle").textContent=t.quotesTitle;
      document.getElementById("articlesEyebrow").textContent=t.articlesEyebrow;
      document.getElementById("articlesTitle").textContent=t.articlesTitle;
      document.getElementById("contactEyebrow").textContent=t.contactEyebrow;
      document.getElementById("contactTitle").textContent=t.contactTitle;
      document.getElementById("contactText").textContent=t.contactText;
      document.getElementById("contactAddressLabel").textContent=t.contactAddressLabel;
      document.getElementById("contactPhoneLabel").textContent=t.contactPhoneLabel;
      document.getElementById("contactHoursLabel").textContent=t.contactHoursLabel;
      document.getElementById("contactFieldLabel").textContent=t.contactFieldLabel;
      document.getElementById("contactFieldText").textContent=t.contactFieldText;
      document.getElementById("phoneLinkContact").textContent=t.phoneContact;
      document.getElementById("zaloLinkContact").textContent=t.zaloContact;
      document.getElementById("contactFormEyebrow").textContent=t.contactFormEyebrow;
      document.getElementById("contactFormTitle").textContent=t.contactFormTitle;
      document.getElementById("contactName").placeholder=t.contactName;
      document.getElementById("contactPhoneInput").placeholder=t.contactPhoneInput;
      document.getElementById("contactMessage").placeholder=t.contactMessage;
      document.getElementById("contactSubmit").textContent=t.contactSubmit;
      document.getElementById("footerContactTitle").textContent=t.footerContactTitle;
      document.getElementById("footerFieldTitle").textContent=t.footerFieldTitle;
      document.getElementById("footerField1").textContent=t.footerField1;
      document.getElementById("footerField2").textContent=t.footerField2;
      document.getElementById("footerField3").textContent=t.footerField3;
      document.getElementById("footerField4").textContent=t.footerField4;
      document.getElementById("floatingTop").textContent=t.floatingTop;
      document.querySelector(".topbar-inner div:last-child").childNodes[0].textContent=`${t.topbarLabel} `;
      document.getElementById("langVi").classList.toggle("active",currentLang==="vi");
      document.getElementById("langEn").classList.toggle("active",currentLang==="en");
    }

    function renderCompany(){
      const company=getContent().company;
      const bullet=currentLang==="vi"?" • ":" • ";
      document.getElementById("companyName").textContent=company.short_name;
      document.getElementById("companyTagline").textContent=company.tagline;
      document.getElementById("companyHours").textContent=company.hours;
      document.getElementById("companyAddress").textContent=company.address;
      document.getElementById("companyPhoneTop").textContent=company.phone;
      document.getElementById("websiteNote").textContent=company.website_note;
      document.getElementById("companyPhoneBox").textContent=company.phone;
      document.getElementById("contactAddress").textContent=company.address;
      document.getElementById("contactPhone").textContent=company.phone;
      document.getElementById("contactZalo").textContent=company.zalo;
      document.getElementById("contactHours").textContent=company.hours;
      document.getElementById("footerName").textContent=company.name;
      document.getElementById("footerTagline").textContent=company.tagline;
      document.getElementById("footerAddress").textContent=company.address;
      document.getElementById("footerPhone").textContent=`${currentLang==="vi"?"SDT":"Phone"}: ${company.phone}`;
      document.getElementById("footerZalo").textContent=`Zalo: ${company.zalo}`;
      document.getElementById("footerHours").textContent=company.hours;
      document.querySelector(".topbar-inner div:first-child").innerHTML=`<span id="companyHours">${company.hours}</span>${bullet}<span id="companyAddress">${company.address}</span>`;
      ["phoneLinkHeader","phoneLinkContact"].forEach(id=>document.getElementById(id).href=phoneHref);
      ["zaloLinkHeader","zaloLinkContact","zaloFloating"].forEach(id=>document.getElementById(id).href=zaloHref);
    }

    function renderHero(){
      const content=getContent();
      const t=pageText[currentLang];
      const slides=document.getElementById("heroSlides");
      const dots=document.getElementById("heroDots");
      clearInterval(slideTimer);
      slides.innerHTML=content.hero_slides.map((slide,index)=>`<article class="slide ${index===slideIndex?"active":""}"><img src="${slide.image}" alt="${slide.title}" /><div class="slide-content"><span class="eyebrow" style="background:rgba(255,255,255,.18);color:#fff;border-color:rgba(255,255,255,.18)">${slide.eyebrow}</span><h1 class="section-title" style="color:white;margin-top:16px;">${slide.title}</h1><p>${slide.description}</p><div style="display:flex;gap:12px;flex-wrap:wrap;"><a class="btn" href="#contact">${t.heroPrimary}</a><a class="btn-outline" style="background:rgba(255,255,255,.16);color:#fff;border-color:rgba(255,255,255,.18)" href="#services">${t.heroSecondary}</a></div></div></article>`).join("");
      dots.innerHTML=content.hero_slides.map((_,index)=>`<button class="dot ${index===slideIndex?"active":""}" data-dot="${index}" type="button"></button>`).join("");
      dots.querySelectorAll("[data-dot]").forEach(button=>button.addEventListener("click",()=>showSlide(Number(button.dataset.dot))));
      slideTimer=setInterval(()=>showSlide((slideIndex+1)%content.hero_slides.length),5000);
    }

    function showSlide(index){
      slideIndex=index;
      document.querySelectorAll(".slide").forEach((slide,i)=>slide.classList.toggle("active",i===index));
      document.querySelectorAll(".dot").forEach((dot,i)=>dot.classList.toggle("active",i===index));
    }

    function renderHighlights(){
      document.getElementById("highlightGrid").innerHTML=getContent().highlights.map(item=>`<article class="card"><strong style="font-size:20px;margin-bottom:8px;">${item.title}</strong><div class="section-text">${item.text}</div></article>`).join("");
    }

    function renderServices(){
      document.getElementById("servicesGrid").innerHTML=getContent().services.map(item=>`<article class="service-card"><img src="${item.image}" alt="${item.name}" /><div class="service-body"><strong style="font-size:20px;">${item.name}</strong><p class="section-text" style="margin-top:8px;">${item.summary}</p></div></article>`).join("");
    }

    function renderTypes(){
      document.getElementById("typeGrid").innerHTML=getContent().project_types.map(item=>`<article class="type-card"><img src="${item.image}" alt="${item.name}" /><div class="type-overlay"><strong>${item.name}</strong><span>${item.style}</span></div></article>`).join("");
    }

    function renderProjects(){
      document.getElementById("projectGrid").innerHTML=getContent().projects.map(item=>`<article class="project-card"><img src="${item.image}" alt="${item.name}" /><div class="project-body"><span class="eyebrow" style="padding:8px 12px;">${item.tag}</span><strong style="font-size:22px;margin:14px 0 8px;">${item.name}</strong><p class="section-text">${item.text}</p></div></article>`).join("");
    }

    function renderProcess(){
      document.getElementById("processGrid").innerHTML=getContent().process.map(item=>`<article class="process-step"><em>${item.step}</em><strong style="font-size:20px;margin:16px 0 8px;">${item.title}</strong><div class="section-text">${item.text}</div></article>`).join("");
    }

    function renderQuotes(){
      document.getElementById("quoteGrid").innerHTML=getContent().testimonials.map(item=>`<article class="quote-card"><strong style="font-size:20px;">${item.name}</strong><span class="article-date" style="margin-top:8px;">${item.role}</span><p class="section-text">${item.quote}</p></article>`).join("");
    }

    function renderArticles(){
      document.getElementById("articleGrid").innerHTML=getContent().articles.map(item=>`<article class="article-card"><img src="${item.image}" alt="${item.title}" /><span class="article-date">${item.date}</span><strong style="font-size:20px;">${item.title}</strong><p class="section-text" style="margin-top:8px;">${item.excerpt}</p></article>`).join("");
    }

    function bindForms(){
      const consultForm=document.getElementById("consultForm");
      const contactForm=document.getElementById("contactForm");
      consultForm.addEventListener("submit",async event=>{
        event.preventDefault();
        const name=document.getElementById("consultName").value.trim();
        const phone=document.getElementById("consultPhone").value.trim();
        const message=document.getElementById("consultMessage").value.trim();
        const status=document.getElementById("consultStatus");
        if(!name||!phone){status.textContent=pageText[currentLang].consultErr;status.className="status error";return;}
        try{
          const response=await fetch("/api/contact",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({form_type:"consult",name,phone,message,language:currentLang,source:"hero_consult_form"})});
          const result=await response.json();
          if(!response.ok){throw new Error(result.error||"Submit failed");}
          status.textContent=pageText[currentLang].consultOk(name,phone);
          status.className="status";
          consultForm.reset();
        }catch(_error){
          status.textContent=currentLang==="vi"?"Không gửi được thông tin. Vui lòng mở website bằng file Open Website.bat hoặc chạy webapp.py.":"Could not submit the information. Please open the website using Open Website.bat or run webapp.py.";
          status.className="status error";
        }
      });
      contactForm.addEventListener("submit",async event=>{
        event.preventDefault();
        const name=document.getElementById("contactName").value.trim();
        const phone=document.getElementById("contactPhoneInput").value.trim();
        const message=document.getElementById("contactMessage").value.trim();
        const status=document.getElementById("contactStatus");
        if(!name||!phone){status.textContent=pageText[currentLang].contactErr;status.className="status error";return;}
        try{
          const response=await fetch("/api/contact",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({form_type:"contact",name,phone,message,language:currentLang,source:"contact_form"})});
          const result=await response.json();
          if(!response.ok){throw new Error(result.error||"Submit failed");}
          status.textContent=pageText[currentLang].contactOk(name);
          status.className="status";
          contactForm.reset();
        }catch(_error){
          status.textContent=currentLang==="vi"?"Không gửi được thông tin. Vui lòng mở website bằng file Open Website.bat hoặc chạy webapp.py.":"Could not submit the information. Please open the website using Open Website.bat or run webapp.py.";
          status.className="status error";
        }
      });
    }

    function setLanguage(lang){
      currentLang=lang;
      localStorage.setItem("thecohome_lang", lang);
      applyStaticText();
      renderCompany();
      renderHero();
      renderHighlights();
      renderServices();
      renderTypes();
      renderProjects();
      renderProcess();
      renderQuotes();
      renderArticles();
    }

    function bindLanguageSwitch(){
      document.getElementById("langVi").addEventListener("click",()=>setLanguage("vi"));
      document.getElementById("langEn").addEventListener("click",()=>setLanguage("en"));
    }

    function bindShareButtons(){
      const headerActions=document.querySelector(".header-actions");
      const contactActions=document.querySelector(".contact-actions");
      const contactCard=document.querySelector("#contact .card");
      const status=document.getElementById("shareStatus") || (() => {
        const node=document.createElement("div");
        node.id="shareStatus";
        node.className="share-feedback";
        if(contactCard){contactCard.appendChild(node);}
        return node;
      })();

      const headerButton=document.createElement("button");
      headerButton.type="button";
      headerButton.id="shareButton";
      headerButton.className="share-mini";
      headerActions?.insertBefore(headerButton, document.getElementById("zaloLinkHeader"));

      const contactButton=document.createElement("button");
      contactButton.type="button";
      contactButton.id="copyLinkButton";
      contactButton.className="share-mini";

      const shareWrap=document.createElement("div");
      shareWrap.className="share-cluster";
      shareWrap.id="shareCluster";
      contactActions?.insertBefore(shareWrap, document.getElementById("phoneLinkContact"));
      shareWrap.appendChild(contactButton);

      const zaloShare=document.createElement("a");
      zaloShare.id="shareZalo";
      zaloShare.className="share-mini zalo";
      zaloShare.target="_blank";
      zaloShare.rel="noreferrer";
      shareWrap.appendChild(zaloShare);

      const facebookShare=document.createElement("a");
      facebookShare.id="shareFacebook";
      facebookShare.className="share-mini facebook";
      facebookShare.target="_blank";
      facebookShare.rel="noreferrer";
      shareWrap.appendChild(facebookShare);

      const messengerShare=document.createElement("a");
      messengerShare.id="shareMessenger";
      messengerShare.className="share-mini messenger";
      messengerShare.target="_blank";
      messengerShare.rel="noreferrer";
      shareWrap.appendChild(messengerShare);

      function iconShare(){
        return '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M4 12v7a1 1 0 0 0 1 1h14a1 1 0 0 0 1-1v-7"/><path d="M12 16V3"/><path d="m7 8 5-5 5 5"/></svg>';
      }
      function iconCopy(){
        return '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="9" y="9" width="11" height="11" rx="2"/><path d="M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1"/></svg>';
      }
      function iconFacebook(){
        return '<svg viewBox="0 0 24 24" fill="currentColor"><path d="M13.5 21v-7h2.4l.4-3h-2.8V9.2c0-.9.3-1.5 1.6-1.5H16V5.1c-.3 0-1-.1-1.9-.1-1.9 0-3.2 1.1-3.2 3.4V11H8.5v3h2.4v7h2.6z"/></svg>';
      }
      function iconMessenger(){
        return '<svg viewBox="0 0 24 24" fill="currentColor"><path d="M12 2C6.5 2 2 6.1 2 11.2c0 2.9 1.5 5.4 3.9 7v3.6l3.4-1.9c.9.2 1.8.3 2.7.3 5.5 0 10-4.1 10-9.2S17.5 2 12 2zm1 12.4-2.5-2.7-4.8 2.7 5.3-5.6 2.5 2.7 4.8-2.7-5.3 5.6z"/></svg>';
      }
      function iconZalo(){
        return '<svg viewBox="0 0 24 24" fill="currentColor"><path d="M4 4h16v16H4z" opacity=".08"/><path d="M7 7h10v2l-6.3 6H17v2H7v-2l6.3-6H7z"/></svg>';
      }

      async function shareLink(){
        const url=window.location.href;
        const title=document.title;
        try{
          if(navigator.share){
            await navigator.share({title,url});
            status.textContent=currentLang==="vi"?"Đã mở bảng chia sẻ.":"Share sheet opened.";
            return;
          }
          await navigator.clipboard.writeText(url);
          status.textContent=currentLang==="vi"?"Đã copy đường dẫn website.":"Website link copied.";
        }catch(_error){
          status.textContent=currentLang==="vi"?"Không thể chia sẻ hoặc copy link trên trình duyệt này.":"Could not share or copy the link on this browser.";
        }
      }

      headerButton.addEventListener("click", shareLink);
      contactButton.addEventListener("click", shareLink);

      function updateShareLinks(){
        const url=encodeURIComponent(window.location.href);
        const text=encodeURIComponent(document.title);
        headerButton.innerHTML=`${iconShare()}<span>${currentLang==="vi"?"Chia sẻ":"Share"}</span>`;
        contactButton.innerHTML=`${iconCopy()}<span>${currentLang==="vi"?"Copy link":"Copy link"}</span>`;
        zaloShare.innerHTML=`${iconZalo()}<span>Zalo</span>`;
        facebookShare.innerHTML=`${iconFacebook()}<span>Facebook</span>`;
        messengerShare.innerHTML=`${iconMessenger()}<span>Messenger</span>`;
        zaloShare.href=`https://zalo.me/share?url=${url}`;
        facebookShare.href=`https://www.facebook.com/sharer/sharer.php?u=${url}`;
        messengerShare.href=`https://www.facebook.com/dialog/send?link=${url}&app_id=291494419107518&redirect_uri=${url}`;
      }

      updateShareLinks();
      window.addEventListener("hashchange", updateShareLinks);
    }

    function bootstrap(){
      if(!["vi","en"].includes(currentLang)){currentLang="vi";}
      applyStaticText();
      renderCompany();
      renderHero();
      renderHighlights();
      renderServices();
      renderTypes();
      renderProjects();
      renderProcess();
      renderQuotes();
      renderArticles();
      bindForms();
      bindLanguageSwitch();
      bindShareButtons();
    }

    bootstrap();
  </script>
</body>
</html>
"""


class StorefrontHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/":
            html = HTML_PAGE.replace("__STORE_DATA__", json.dumps(STORE_DATA, ensure_ascii=False))
            self._send_html(html)
            return
        if parsed.path == "/api/storefront":
            self._send_json(STORE_DATA)
            return
        self._send_json({"error": "Not found"}, status=HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path != "/api/contact":
            self._send_json({"error": "Not found"}, status=HTTPStatus.NOT_FOUND)
            return
        content_length = int(self.headers.get("Content-Length", "0"))
        raw_body = self.rfile.read(content_length)
        try:
            payload = json.loads(raw_body.decode("utf-8"))
        except json.JSONDecodeError:
            self._send_json({"error": "Invalid JSON payload."}, status=HTTPStatus.BAD_REQUEST)
            return

        name = str(payload.get("name", "")).strip()
        phone = str(payload.get("phone", "")).strip()
        if not name or not phone:
            self._send_json({"error": "Name and phone are required."}, status=HTTPStatus.BAD_REQUEST)
            return

        save_contact(
            {
                "form_type": str(payload.get("form_type", "")).strip(),
                "name": name,
                "phone": phone,
                "message": str(payload.get("message", "")).strip(),
                "language": str(payload.get("language", "")).strip(),
                "source": str(payload.get("source", "")).strip(),
            }
        )
        self._send_json({"status": "ok"}, status=HTTPStatus.CREATED)

    def log_message(self, format: str, *args) -> None:
        return

    def _send_html(self, html: str, status: HTTPStatus = HTTPStatus.OK) -> None:
        data = html.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _send_json(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def main() -> None:
    ensure_contacts_workbook()
    server = ThreadingHTTPServer((HOST, PORT), StorefrontHandler)
    print(f"Website running at http://{HOST}:{PORT}")
    threading.Timer(1.0, lambda: webbrowser.open(f"http://{HOST}:{PORT}")).start()
    server.serve_forever()


if __name__ == "__main__":
    main()
